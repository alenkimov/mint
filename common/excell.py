from typing import Any, List, Dict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


class Column:
    DESCRIPTION_CELL_COLOR_IF_REQUIRED = "90EE90"

    def __init__(
            self,
            header: str,
            description: str,
            key: str,
            *,
            group_name: str = None,
            required: bool = False,
            max_length: int = None
    ) -> None:
        self.header = header
        self.description = description
        self.key = key
        self.group_name = group_name
        self.required = required
        self.max_length: int = max_length

    @property
    def color(self):
        if self.required:
            return PatternFill(
                start_color=self.DESCRIPTION_CELL_COLOR_IF_REQUIRED,
                end_color=self.DESCRIPTION_CELL_COLOR_IF_REQUIRED,
                fill_type="solid",
            )

        return None

    @property
    def full_header(self) -> str:
        header = f"[{self.group_name}] " if self.group_name else ""
        header += self.header
        return header

    @property
    def full_description(self) -> str:
        description = "REQUIRED FIELD" if self.required else "OPTIONAL FIELD"
        if self.max_length:
            description += f"\nMAX LENGTH: {self.max_length}"
        description += f"\n\n{self.description}"
        return description


class Excel:
    def __init__(self, columns: list[Column]) -> None:
        self.columns = columns

    def create_empty_table(self, dirpath: Path, name: str):
        """Создает пустую таблицу с заголовками и описаниями."""
        wb = Workbook()
        ws = wb.active

        for i, column in enumerate(self.columns, start=1):
            ws.cell(row=1, column=i, value=column.full_header)

        for i, column in enumerate(self.columns, start=1):
            cell = ws.cell(row=2, column=i, value=column.full_description)

            if column.color:
                cell.fill = column.color

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length

        table_filepath = dirpath / f"{name}.xlsx"
        wb.save(table_filepath)

    def read_table(self, filepath: Path) -> List[Dict[str, Any]]:
        """Читает данные из таблицы начиная с третьей строки."""
        wb = load_workbook(filepath)
        ws = wb.active

        data = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            row_data = {}
            for i, value in enumerate(row):
                column = self.columns[i]
                if column.group_name:
                    if column.group_name not in row_data:
                        row_data[column.group_name] = {}
                    row_data[column.group_name][column.key] = value
                else:
                    row_data[column.key] = value
            data.append(row_data)

        return data


def get_xlsx_filepathes(dirpath: Path) -> list[Path]:
    return list(dirpath.glob("*.xlsx"))
